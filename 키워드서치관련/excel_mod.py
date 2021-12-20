import zipfile
from collections import OrderedDict
import pandas as pd
import xlrd
import zipfile
import re
from openpyxl import load_workbook
from timeit import default_timer as timer
from datetime import timedelta
from xml.etree.ElementTree import iterparse
try:
    from xml.etree.cElementTree import XML
except ImportError:
    from xml.etree.ElementTree import XML


def test_xlrd(full_path):
    workbook = xlrd.open_workbook(full_path, on_demand=True)
    for sheet in workbook.sheet_names():
        val_list = []
        worksheet = workbook.sheet_by_name(sheet)
        nrows = worksheet.nrows
        ncols = worksheet.ncols
        for row_num in range(nrows):
            for col_num in range(ncols):
                value = worksheet.cell_value(row_num, col_num)
                val_list.append(value)
        value_list = list(OrderedDict.fromkeys(val_list))
        txt = ','.join(value_list)
        print(txt)

def openpyxl(full_path):
    load_wb = load_workbook(full_path, data_only=True)
    for sheet in load_wb.sheetnames:
        txt = ''
        load_ws = load_wb[sheet]
        for row in load_ws.rows:
            for cell in row:
                if cell.value == None:
                    continue
                txt += str(cell.value)
        print(txt)

def xml(full_path):
    z = zipfile.ZipFile(full_path)
    file_name_list = []
    for f in z.filelist:
        file_name_list.append(f)
    for file_name in file_name_list:
        if file_name.filename == 'xl/sharedStrings.xml':
            strings = [el.text for e, el in iterparse(z.open('xl/sharedStrings.xml')) if
                       el.tag.endswith('}t')]
            nums = []
            for d in z.namelist():
                if d.startswith("xl/worksheets/sheet"):
                    nums.append(int(d[len("xl/worksheets/sheet"):-4]))
            s_format = "xl/worksheets/sheet%s.xml"
            sheet_name_list = [s_format % x for x in sorted(nums)]
            value = ''
            for sheet in sheet_name_list:
                value_list = []
                for e, el in iterparse(z.open(sheet)):
                    if el.tag.endswith('}v'):
                        value = el.text
                    if el.tag.endswith('}c'):
                        if el.attrib.get('t') == 's':
                            value = str(strings[int(value)]).upper() + '\n'
                            value_list.append(value)
                value_list = list(OrderedDict.fromkeys(value_list))
                txt = ','.join(value_list)
                print(txt)

def pandas(full_path):
    dataframe = pd.read_excel(full_path, engine="openpyxl", sheet_name=None)
    pd.set_option('display.max.colwidth', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    for col_name, df in dataframe.items():
        val_list = df.values.tolist()
        for value in val_list:
            value_list = list(OrderedDict.fromkeys(str(value)))
            txt = ','.join(value_list) + '\n'
            print(txt)

def xml_xlsx_2021(fname, sheet):
    z = zipfile.ZipFile(fname)
    if 'xl/sharedStrings.xml' in z.namelist():
        # Get shared strings
        strings = [element.text for event, element
                   in iterparse(z.open('xl/sharedStrings.xml'))
                   if element.tag.endswith('}t')]
    sheetdict = {element.attrib['name']: element.attrib['sheetId'] for event, element in
                 iterparse(z.open('xl/workbook.xml'))
                 if element.tag.endswith('}sheet')}
    rows = []
    row = {}
    value = ''

    if sheet in sheets:
        sheetfile = 'xl/worksheets/sheet' + sheets[sheet] + '.xml'
    # print(sheet,sheetfile)
    for event, element in iterparse(z.open(sheetfile)):
        # get value or index to shared strings
        if element.tag.endswith('}v') or element.tag.endswith('}t'):
            value = element.text
        # If value is a shared string, use value as an index
        if element.tag.endswith('}c'):
            if element.attrib.get('t') == 's':
                value = strings[int(value)]
            # split the row/col information so that the row leter(s) can be separate
            letter = re.sub('\d', '', element.attrib['r'])
            row[letter] = value
            value = ''
        if element.tag.endswith('}row'):
            rows.append(row)
            row = {}

    return rows


if __name__ == '__main__':
    start = timer()
    full_path = r''
    xml(full_path)

    end = timer()
    print(timedelta(seconds=end - start))